"""
Importador de NF-e (DANFE) → Excel
===================================
100% gratuito, roda offline, sem API.

Requisitos (instalar uma vez):
  pip install pypdf openpyxl

Uso:
  python importar_nfe.py              (detecta PDF automaticamente na pasta)
  python importar_nfe.py nota.pdf     (especifica o arquivo)
"""

import re
import sys
import argparse
from pathlib import Path

try:
    from pypdf import PdfReader
except ImportError:
    print("❌ Instale o pypdf:  pip install pypdf")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Instale o openpyxl:  pip install openpyxl")
    sys.exit(1)


# ─────────────────────────────────────────────
#  EXTRAÇÃO DE TEXTO DO PDF
# ─────────────────────────────────────────────

def extrair_texto_pdf(caminho_pdf: str) -> str:
    reader = PdfReader(caminho_pdf)
    partes = []
    for page in reader.pages:
        t = page.extract_text()
        if t:
            partes.append(t)
    return "\n".join(partes)


# ─────────────────────────────────────────────
#  PARSE DO CABEÇALHO
# ─────────────────────────────────────────────

def parsear_cabecalho(texto: str) -> dict:
    cab = {}

    m = re.search(r'Nº\s+([\d.]+)', texto)
    cab['numero_nfe'] = m.group(1).strip() if m else ''

    m = re.search(r'SÉRIE\s+(\d+)', texto)
    cab['serie'] = m.group(1) if m else ''

    m = re.search(r'DATA DA EMISSÃO\s+(\d{2}/\d{2}/\d{4})', texto)
    cab['data_emissao'] = m.group(1) if m else ''

    m = re.search(r'DATA DA SAÍDA\s+(\d{2}/\d{2}/\d{4})', texto)
    cab['data_saida'] = m.group(1) if m else ''

    # Emitente (linha após DANFE)
    m = re.search(r'DANFE(.+?)\n', texto)
    cab['emitente_nome'] = m.group(1).strip() if m else ''

    # CNPJ emitente
    m = re.search(r'INSCRIÇÃO ESTADUAL DO SUBSTITUTO TRIBUTÁRIO CNPJ\s+([\d./\-]+)', texto)
    cab['emitente_cnpj'] = m.group(1).strip() if m else ''

    # Destinatário
    m = re.search(r'NOME / RAZÃO SOCIAL\s+(.+?)\nDATA DA EMISSÃO', texto)
    cab['destinatario_nome'] = m.group(1).strip() if m else ''

    m = re.search(r'CNPJ / CPF\s+([\d./\-]+)', texto)
    cab['destinatario_cnpj'] = m.group(1).strip() if m else ''

    m = re.search(r'MUNICÍPIO\s+(\w[\w\s]+?)\n', texto)
    cab['municipio'] = m.group(1).strip() if m else ''

    m = re.search(r'Valor Total:\s*([\d.,]+)', texto)
    cab['valor_total'] = m.group(1) if m else ''

    m = re.search(r'(\d{4} \d{4} \d{4} \d{4} \d{4} \d{4} \d{4} \d{4} \d{4} \d{4} \d{4})', texto)
    cab['chave_acesso'] = m.group(1).strip() if m else ''

    m = re.search(r'NATUREZA DA OPERAÇÃO\s+(.+?)\n', texto)
    cab['natureza_operacao'] = m.group(1).strip() if m else ''

    m = re.search(r'Vencimento\s*\n:\s*\n(\d{2}/\d{2}/\d{4})', texto)
    cab['vencimento_duplicata'] = m.group(1) if m else ''

    return cab


# ─────────────────────────────────────────────
#  PARSE DOS ITENS
# ─────────────────────────────────────────────

def br_to_float(s: str) -> float:
    if not s:
        return 0.0
    s = s.strip().replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return 0.0


def parsear_itens(texto: str) -> list:
    """
    Com pypdf o texto vem diferente: a linha de dados numéricos
    (NCM CST CFOP UNID QTDE V_UNIT V_DESC V_LIQ ...) fica APÓS
    o bloco de descrição/lote. Estratégia: capturar bloco por bloco.
    """
    itens = []

    # Cada item começa com código como 85/1 ou 207/1
    # Bloco típico (pypdf):
    #   211/1 TRESBOMM - ALOJAMENTO C/ AMOXI E TIAMU\nVACCINAR\n
    #   Lote: 5003 Data Fab.: 17/02/2026 Vcto: 19/03/2026 Fab: 3BS\n...
    #   23099010 050 5124 KG 18.032,20 0,12 0,00 2.163,86 ...

    # Regex para linha de dados numéricos ao final do bloco de cada item
    dados_re = re.compile(
        r'(\d{8})\s+'          # NCM
        r'(\d{3})\s+'          # CST
        r'(\d{4})\s+'          # CFOP
        r'(KG|UN|PC|CX|SC|LT)\s+'  # unidade
        r'([\d.]+,\d+)\s+'     # qtde
        r'([\d.]+,\d+)\s+'     # valor unitário
        r'([\d.]+,\d+)\s+'     # valor desconto
        r'([\d.]+,\d+)'        # valor líquido
    )

    lote_re = re.compile(r'Lote:\s*(\S+)')
    fab_re  = re.compile(r'Data Fab\.:\s*(\d{2}/\d{2}/\d{4})')
    val_re  = re.compile(r'Vcto:\s*(\d{2}/\d{2}/\d{4})')
    cod_re  = re.compile(r'^(\d+/\d+)\s+(.+)')

    linhas = texto.split('\n')
    i = 0

    while i < len(linhas):
        linha = linhas[i].strip()

        # Verifica se é linha de início de item (código + descrição)
        mc = cod_re.match(linha)
        if mc:
            codigo = mc.group(1)
            descricao_partes = [mc.group(2).strip()]

            # Coletar bloco até encontrar a linha de dados numéricos
            bloco_linhas = [linha]
            j = i + 1
            dados_match = None

            while j < len(linhas) and j < i + 15:
                prox = linhas[j].strip()
                bloco_linhas.append(prox)
                md = dados_re.search(prox)
                if md:
                    dados_match = md
                    # descrição é tudo antes dos dados numéricos
                    # (linhas entre o código e a linha de NCM)
                    break
                # Continua acumulando descrição se não for linha de lote/fab/ncm
                if prox and not re.match(r'Lote:|FAB:|VAL:|QTD:|LOTE:|23\d{6}', prox):
                    descricao_partes.append(prox)
                j += 1

            if dados_match:
                bloco_texto = '\n'.join(bloco_linhas)

                # Limpa descrição: remove partes que são lote/fab/etc
                desc_limpa = ' '.join(
                    p for p in descricao_partes
                    if p and not re.match(r'Lote:|FAB:|VAL:|QTD:|LOTE:|ANIMAL FEED|23\d{6}|VACCINAR$|CARGILL$', p)
                ).strip()

                lote = lote_re.search(bloco_texto)
                fab  = fab_re.search(bloco_texto)
                val  = val_re.search(bloco_texto)

                itens.append({
                    'codigo':          codigo,
                    'descricao':       desc_limpa,
                    'lote':            lote.group(1) if lote else '',
                    'data_fabricacao': fab.group(1)  if fab  else '',
                    'data_validade':   val.group(1)  if val  else '',
                    'unidade':         dados_match.group(4),
                    'quantidade':      br_to_float(dados_match.group(5)),
                    'valor_unitario':  br_to_float(dados_match.group(6)),
                    'valor_desconto':  br_to_float(dados_match.group(7)),
                    'valor_liquido':   br_to_float(dados_match.group(8)),
                })
                i = j + 1
                continue

        i += 1

    return itens


# ─────────────────────────────────────────────
#  GERAÇÃO DO EXCEL
# ─────────────────────────────────────────────

AZUL_ESC = "1F4E79"
AZUL_CLR = "D6E4F0"
BRANCO   = "FFFFFF"

thin  = Side(style='thin', color='AAAAAA')
borda = Border(left=thin, right=thin, top=thin, bottom=thin)


def aplicar_borda(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = borda


def celula_cabec(ws, row, col, valor):
    c = ws.cell(row=row, column=col, value=valor)
    c.font = Font(bold=True, color=BRANCO, name='Arial', size=10)
    c.fill = PatternFill('solid', start_color=AZUL_ESC)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    return c


def gerar_excel(cabecalho: dict, itens: list, caminho_saida: str):
    wb = openpyxl.Workbook()

    # ── ABA 1: CABEÇALHO ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Cabeçalho NF-e'

    ws1.merge_cells('A1:B1')
    t = ws1['A1']
    t.value = f"NOTA FISCAL ELETRÔNICA – NF-e Nº {cabecalho.get('numero_nfe', '')}"
    t.font = Font(bold=True, name='Arial', size=13, color=BRANCO)
    t.fill = PatternFill('solid', start_color=AZUL_ESC)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 30

    campos = [
        ('NF-e Número',           cabecalho.get('numero_nfe', '')),
        ('Série',                  cabecalho.get('serie', '')),
        ('Data Emissão',           cabecalho.get('data_emissao', '')),
        ('Data Saída',             cabecalho.get('data_saida', '')),
        ('Natureza da Operação',   cabecalho.get('natureza_operacao', '')),
        ('Emitente',               cabecalho.get('emitente_nome', '')),
        ('CNPJ Emitente',          cabecalho.get('emitente_cnpj', '')),
        ('Destinatário',           cabecalho.get('destinatario_nome', '')),
        ('CNPJ Destinatário',      cabecalho.get('destinatario_cnpj', '')),
        ('Município',              cabecalho.get('municipio', '')),
        ('Valor Total da NF-e',    cabecalho.get('valor_total', '')),
        ('Vencimento Duplicata',   cabecalho.get('vencimento_duplicata', '')),
        ('Chave de Acesso',        cabecalho.get('chave_acesso', '')),
    ]

    for i, (campo, valor) in enumerate(campos, start=2):
        bg = AZUL_CLR if i % 2 == 0 else BRANCO
        c1 = ws1.cell(row=i, column=1, value=campo)
        c1.font  = Font(bold=True, name='Arial', size=10, color=BRANCO)
        c1.fill  = PatternFill('solid', start_color=AZUL_ESC)
        c1.alignment = Alignment(vertical='center')

        c2 = ws1.cell(row=i, column=2, value=valor)
        c2.font = Font(name='Arial', size=10)
        c2.fill = PatternFill('solid', start_color=bg)
        c2.alignment = Alignment(vertical='center', wrap_text=True)
        ws1.row_dimensions[i].height = 18

    ws1.column_dimensions['A'].width = 28
    ws1.column_dimensions['B'].width = 72
    aplicar_borda(ws1, 1, len(campos) + 1, 1, 2)

    # ── ABA 2: ITENS ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Itens NF-e')

    colunas = ['#', 'Código', 'Descrição do Produto', 'Lote',
               'Data Fabricação', 'Data Validade', 'Unid.',
               'Quantidade', 'Valor Unit. (R$)', 'Valor Desconto (R$)', 'Valor Líquido (R$)']

    for col, h in enumerate(colunas, 1):
        celula_cabec(ws2, 1, col, h)
    ws2.row_dimensions[1].height = 36

    total_qtde = 0.0
    total_vliq = 0.0

    for idx, item in enumerate(itens, 1):
        row = idx + 1
        bg  = AZUL_CLR if idx % 2 == 0 else BRANCO

        valores = [
            idx,
            item['codigo'],
            item['descricao'],
            item['lote'],
            item['data_fabricacao'],
            item['data_validade'],
            item['unidade'],
            item['quantidade'],
            item['valor_unitario'],
            item['valor_desconto'],
            item['valor_liquido'],
        ]

        for col, val in enumerate(valores, 1):
            c = ws2.cell(row=row, column=col, value=val)
            c.font = Font(name='Arial', size=10)
            c.fill = PatternFill('solid', start_color=bg)
            c.border = borda
            c.alignment = Alignment(
                vertical='center',
                horizontal='right'  if col >= 8  else
                           'center' if col in (1, 7) else 'left',
                wrap_text=(col == 3)
            )

        ws2.cell(row=row, column=8).number_format  = '#,##0.00'
        ws2.cell(row=row, column=9).number_format  = 'R$ #,##0.00000'
        ws2.cell(row=row, column=10).number_format = 'R$ #,##0.00'
        ws2.cell(row=row, column=11).number_format = 'R$ #,##0.00'

        total_qtde += item['quantidade']
        total_vliq += item['valor_liquido']

    # Linha de total
    total_row = len(itens) + 2
    ws2.merge_cells(f'A{total_row}:G{total_row}')
    ct = ws2.cell(row=total_row, column=1, value='TOTAL GERAL')
    ct.font = Font(bold=True, name='Arial', size=11, color=BRANCO)
    ct.fill = PatternFill('solid', start_color=AZUL_ESC)
    ct.alignment = Alignment(horizontal='center', vertical='center')

    for col, val, fmt in [
        (8,  total_qtde, '#,##0.00'),
        (10, 0.00,       'R$ #,##0.00'),
        (11, total_vliq, 'R$ #,##0.00'),
    ]:
        c = ws2.cell(row=total_row, column=col, value=val)
        c.font = Font(bold=True, name='Arial', size=11, color=BRANCO)
        c.fill = PatternFill('solid', start_color=AZUL_ESC)
        c.number_format = fmt
        c.border = borda
        c.alignment = Alignment(horizontal='right', vertical='center')

    ws2.cell(row=total_row, column=9).fill   = PatternFill('solid', start_color=AZUL_ESC)
    ws2.cell(row=total_row, column=9).border = borda

    larguras = [5, 10, 52, 9, 16, 16, 7, 16, 18, 20, 20]
    for i, w in enumerate(larguras, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws2.freeze_panes = 'A2'
    aplicar_borda(ws2, 1, total_row, 1, 11)

    wb.save(caminho_saida)
    return total_qtde, total_vliq


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='Converte NF-e (DANFE PDF) em planilha Excel'
    )
    parser.add_argument('pdf', nargs='?', default='',
                        help='Caminho do PDF (opcional – detecta automaticamente se omitido)')
    parser.add_argument('-o', '--output', default='',
                        help='Nome do arquivo Excel de saída (opcional)')
    args = parser.parse_args()

    # Auto-detectar PDF na pasta
    if args.pdf:
        pdf_path = Path(args.pdf)
        if not pdf_path.exists():
            print(f"❌ Arquivo não encontrado: {pdf_path}")
            sys.exit(1)
    else:
        pdfs = list(Path('.').glob('*.pdf')) + list(Path('.').glob('*.PDF'))
        if len(pdfs) == 0:
            print("❌ Nenhum arquivo PDF encontrado na pasta.")
            print("   Coloque o PDF na mesma pasta do script e tente novamente.")
            sys.exit(1)
        elif len(pdfs) == 1:
            pdf_path = pdfs[0]
            print(f"📎 PDF detectado automaticamente: {pdf_path.name}")
        else:
            print("📎 Vários PDFs encontrados. Escolha um:\n")
            for i, p in enumerate(pdfs, 1):
                print(f"   {i}. {p.name}")
            print()
            while True:
                try:
                    escolha = int(input("Digite o número: ")) - 1
                    if 0 <= escolha < len(pdfs):
                        pdf_path = pdfs[escolha]
                        break
                    else:
                        print("   Número inválido, tente novamente.")
                except ValueError:
                    print("   Digite apenas o número.")

    xlsx_path = args.output if args.output else pdf_path.stem + '.xlsx'

    print(f"\n📄 Lendo PDF: {pdf_path.name}")
    texto = extrair_texto_pdf(str(pdf_path))

    print("🔍 Extraindo cabeçalho...")
    cabecalho = parsear_cabecalho(texto)

    print("📦 Extraindo itens de produto...")
    itens = parsear_itens(texto)

    if not itens:
        print("⚠️  Nenhum item encontrado. Verifique se o PDF é um DANFE válido.")
        sys.exit(1)

    print(f"✅ {len(itens)} itens encontrados")
    print(f"💾 Gerando Excel: {xlsx_path}")

    total_qtde, total_vliq = gerar_excel(cabecalho, itens, xlsx_path)

    print(f"\n{'─'*50}")
    print(f"  NF-e Nº:       {cabecalho.get('numero_nfe', '')}")
    print(f"  Emissão:       {cabecalho.get('data_emissao', '')}")
    print(f"  Itens:         {len(itens)}")
    print(f"  Total Qtde:    {total_qtde:,.2f} kg/un")
    print(f"  Valor Total:   R$ {total_vliq:,.2f}")
    print(f"{'─'*50}")
    print(f"\n✅ Arquivo salvo: {xlsx_path}\n")


if __name__ == '__main__':
    main()